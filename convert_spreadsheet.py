# load the spreadsheet and convert to JSON
import openpyxl
import json
import string
import random

def GetRandomID():

	N = 7
	res = ''.join(random.choices(string.ascii_uppercase + string.digits, k=N))
	
	return res


wb = openpyxl.load_workbook('OrderSetBuild.xlsx')
ws = wb.get_sheet_by_name('HighPriority')

row_count = ws.max_row
column_count = ws.max_column

print("Number of rows: " + str(row_count))
print("Number of columns: " + str(column_count))

coldesc = {"order_set": 3, "type": 6, "cst_order": 7, "nh_order": 8, "word_file": 9, "nh_uptodate": 11, "comment": 12, "stakeholders": 13, "dcw_owner": 14, "updated": 15, "current_dcw": 16}
outjson = {"ordersets": []}
tindex = 1
tbundle = 1
for trow in range(3, row_count):

	#for tcol in range(0, column_count):
	rowdata = {}
	for tcol in coldesc:
		tcell = ws.cell(row=trow, column=coldesc[tcol])
		if (tcol != "cst_order") and (tcol != "nh_order") and (tcol != "word_file") and (tcol != "current_dcw"):
			tval = tcell.value
			if tval is None:
				tval = "NA"
			rowdata[tcol] = str(tval)
			if tcol == "updated":
				ldata = str(tval).split(' ')
				tval = ldata[0]
				rowdata[tcol] = tval
		else:
			if tcell.hyperlink is not None:
				rowdata[tcol] = tcell.hyperlink.target
			if tcell.hyperlink is None:
				rowdata[tcol] = 'NA'
				
	# set index and identifier
	rowdata["index"] = str(tindex)
	rowdata["id"] = GetRandomID()
	if (tindex > 105) and (tindex < 201):
		tbundle = 2
	if (tindex > 200) and (tindex < 301):
		tbundle = 3
	rowdata["bundle"] = tbundle
	rowdata["flag"] = "green"
	rowdata["working_group"] = "NA"
	rowdata["service_network"] = "NA"
	rowdata["build_status"] = "DCW"
	rowdata["word_file"] = "NA"
	rowdata["zynx_link"] = "NA"
	rowdata["hls_owner"] = "NA"
	rowdata["version"] = "0.1"
	rowdata["version1_start"] = "2023-03-15"
	rowdata["version1_end"] = "2023-04-15"
	rowdata["version2_start"] = "2023-04-15"
	rowdata["version2_end"] = "2023-05-15"
	rowdata["version3_start"] = "2023-05-15"
	rowdata["version3_end"] = "2023-06-15"
	
	# save it
	outjson["ordersets"].append(rowdata)
	
	tindex += 1

jsondata = json.dumps(outjson)

f = open("order.json", "w")
f.write(jsondata)
f.close()
